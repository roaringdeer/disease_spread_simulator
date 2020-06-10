from copy import deepcopy

from openpyxl import load_workbook


def go():
    wb = load_workbook(filename="graph_definition.xlsx", data_only=True)
    ws = wb["matrix"]
    data_range = "B2:AX50"
    output = [[ws[data_range][i][j].value for j in range(len(ws[data_range][i]))] for i in range(len(ws[data_range]))]
    return output
